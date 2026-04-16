#!/usr/bin/env python3
"""
visualize.py — визуализация результатов анализа журналов LabGenerator TeacherEmulator.

Загружает все файлы analysis.json из указанного каталога (рекурсивно)
и строит графики двух уровней:

  Общий анализ:
  1. quality_heatmap.png          — тепловая карта оценок по всем тестам и ЛР
  2. quality_averages.png         — средние оценки качества по прогонам анализа
  3. boolean_criteria.png         — доля выполнения булевых критериев (все данные)
  4. score_distribution.png       — распределение всех оценок (гистограмма)

  В разрезе дисциплин (строится при наличии двух и более дисциплин):
  5. disc_quality_averages.png    — средние оценки качества по каждой дисциплине
  6. disc_boolean_criteria.png    — тепловая карта критериев по дисциплинам (% «Да»)

  Экспорт в Excel (требует: pip install openpyxl):
  7. analysis.xlsx — пять листов:
       «Качество ЛР»         — оценки по каждой ЛР каждого теста
       «Критерии дисциплин»  — булевы критерии на уровне дисциплин
       «Критерии ЛР»         — булевы критерии и ошибки на уровне ЛР
       «Сводка по дисциплинам» — средние и доли по каждой дисциплине
       «Сводка по прогонам»  — средние оценки по каждому прогону анализа

Использование:
  python visualize.py
  python visualize.py --input path/to/journal-analysis
  python visualize.py --input path/to/analysis.json --output charts/
  python visualize.py --show
  python visualize.py --no-excel
"""

import argparse
import json
import pathlib
import re
import sys
from collections import defaultdict

import matplotlib.pyplot as plt
import numpy as np


QUALITY_METRICS = ["Correctness", "Quality", "Completeness", "Clarity"]

METRIC_LABELS = {
    "Correctness": "Корректность",
    "Quality": "Качество",
    "Completeness": "Полнота",
    "Clarity": "Ясность",
}

BOOLEAN_CRITERIA = [
    ("AssignmentsMatchDiscipline", "Соответствие\nдисциплине"),
    ("LabsDiffer",                 "Различие\nЛР"),
    ("SequenceLogical",            "Логичность\nпоследовательности"),
    ("VariantsDiffer",             "Различие\nвариантов"),
    ("VariantsSameDifficulty",     "Одинаковая\nсложность"),
]

SCORE_COLORS = {
    0: "#c0392b",
    1: "#e74c3c",
    2: "#e67e22",
    3: "#f1c40f",
    4: "#2ecc71",
    5: "#27ae60",
}


# ---------------------------------------------------------------------------
# Загрузка данных
# ---------------------------------------------------------------------------

def find_analysis_files(root: pathlib.Path) -> list[pathlib.Path]:
    if root.is_file():
        return [root] if root.name == "analysis.json" else []
    return sorted(root.rglob("analysis.json"))


def load_all(paths: list[pathlib.Path]) -> list[dict]:
    result = []
    for p in paths:
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            data["_run_name"] = p.parent.name
            result.append(data)
        except Exception as exc:
            print(f"[warn] Пропуск {p}: {exc}", file=sys.stderr)
    return result


def normalize_discipline(name: str) -> str:
    """Убирает суффикс '(Test N)' из имени дисциплины."""
    return re.sub(r"\s*\(Test\s+\d+\)\s*$", "", name, flags=re.IGNORECASE).strip()


def extract_quality_rows(all_data: list[dict]) -> list[dict]:
    rows = []
    for run in all_data:
        run_name = run["_run_name"]
        for result in run.get("Results", []):
            test_id = result.get("RunId", "?")
            disc_name = normalize_discipline(
                result.get("Discipline", {}).get("Name", "Неизвестная дисциплина")
            )
            for q in result.get("Quality", []):
                lab_n = q.get("LabNumber", "?")
                rows.append({
                    "run":        run_name,
                    "test":       test_id,
                    "lab":        lab_n,
                    "discipline": disc_name,
                    "label":      f"{test_id} / ЛР{lab_n}",
                    **{m: q.get(m) for m in QUALITY_METRICS},
                })
    return rows


def extract_boolean_rows(all_data: list[dict]) -> list[dict]:
    rows = []
    for run in all_data:
        for result in run.get("Results", []):
            test_id = result.get("RunId", "?")
            disc = result.get("Discipline", {})
            disc_name = normalize_discipline(disc.get("Name", "Неизвестная дисциплина"))
            rows.append({
                "test":                       test_id,
                "discipline":                 disc_name,
                "AssignmentsMatchDiscipline": disc.get("AssignmentsMatchDiscipline"),
                "LabsDiffer":                 disc.get("LabsDiffer"),
                "SequenceLogical":            disc.get("SequenceLogical"),
                "VariantsDiffer":             None,
                "VariantsSameDifficulty":     None,
            })
            for lab in result.get("Labs", []):
                rows.append({
                    "test":                       test_id,
                    "discipline":                 disc_name,
                    "AssignmentsMatchDiscipline": None,
                    "LabsDiffer":                 None,
                    "SequenceLogical":            None,
                    "VariantsDiffer":             lab.get("VariantsDiffer"),
                    "VariantsSameDifficulty":     lab.get("VariantsSameDifficulty"),
                })
    return rows


# ---------------------------------------------------------------------------
# Графики
# ---------------------------------------------------------------------------

def _save(fig: plt.Figure, out_dir: pathlib.Path, filename: str, show: bool) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / filename
    fig.savefig(path, dpi=150, bbox_inches="tight")
    print(f"  сохранён: {path}")
    if show:
        plt.show()
    plt.close(fig)


def plot_quality_heatmap(
    quality_rows: list[dict],
    out_dir: pathlib.Path,
    show: bool,
    max_rows: int = 60,
) -> None:
    """Тепловая карта оценок: строки = тест/ЛР, столбцы = метрики."""
    if not quality_rows:
        return

    rows = quality_rows[:max_rows]
    if len(quality_rows) > max_rows:
        print(f"  [info] Тепловая карта: показаны первые {max_rows} из {len(quality_rows)} строк.")

    labels = [r["label"] for r in rows]
    data = np.array(
        [[r[m] if r[m] is not None else np.nan for m in QUALITY_METRICS] for r in rows],
        dtype=float,
    )

    fig_h = max(4.0, len(rows) * 0.32 + 1.5)
    fig, ax = plt.subplots(figsize=(7, fig_h))

    im = ax.imshow(data, cmap="RdYlGn", vmin=0, vmax=5, aspect="auto")

    ax.set_xticks(range(len(QUALITY_METRICS)))
    ax.set_xticklabels([METRIC_LABELS[m] for m in QUALITY_METRICS], fontsize=10)
    ax.set_yticks(range(len(rows)))
    ax.set_yticklabels(labels, fontsize=7)
    ax.tick_params(top=True, labeltop=True, bottom=False, labelbottom=False)

    for i in range(len(rows)):
        for j, m in enumerate(QUALITY_METRICS):
            val = data[i, j]
            if not np.isnan(val):
                ax.text(j, i, str(int(val)), ha="center", va="center",
                        fontsize=8, fontweight="bold",
                        color="white" if val < 2 else "black")

    plt.colorbar(im, ax=ax, label="Оценка (0–5)", shrink=0.6)
    ax.set_title("Тепловая карта качества заданий", pad=14)
    fig.tight_layout()
    _save(fig, out_dir, "quality_heatmap.png", show)


def plot_quality_averages(
    quality_rows: list[dict],
    out_dir: pathlib.Path,
    show: bool,
) -> None:
    """Сгруппированные столбики: средние оценки по прогонам."""
    if not quality_rows:
        return

    run_values: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))
    for r in quality_rows:
        for m in QUALITY_METRICS:
            if r[m] is not None:
                run_values[r["run"]][m].append(r[m])

    runs = sorted(run_values)
    x = np.arange(len(runs))
    n_metrics = len(QUALITY_METRICS)
    width = 0.18
    colors = ["#e74c3c", "#3498db", "#2ecc71", "#f39c12"]

    fig, ax = plt.subplots(figsize=(max(6, len(runs) * 1.6 + 1), 5))

    for i, m in enumerate(QUALITY_METRICS):
        means = [
            float(np.mean(run_values[r][m])) if run_values[r][m] else 0.0
            for r in runs
        ]
        offset = (i - (n_metrics - 1) / 2) * width
        bars = ax.bar(x + offset, means, width, label=METRIC_LABELS[m],
                      color=colors[i], zorder=3)
        for bar, val in zip(bars, means):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.06,
                f"{val:.1f}",
                ha="center", va="bottom", fontsize=7,
            )

    ax.set_ylim(0, 5.9)
    ax.axhline(5, color="gray", linestyle="--", linewidth=0.8, alpha=0.5, zorder=0)
    ax.set_ylabel("Средняя оценка")
    ax.set_title("Средние оценки качества по прогонам анализа")
    ax.set_xticks(x)
    ax.set_xticklabels(runs, rotation=20, ha="right", fontsize=8)
    ax.legend(loc="lower right", fontsize=8)
    ax.yaxis.grid(True, linestyle=":", alpha=0.4, zorder=0)
    ax.set_axisbelow(True)
    fig.tight_layout()
    _save(fig, out_dir, "quality_averages.png", show)


def plot_boolean_criteria(
    boolean_rows: list[dict],
    out_dir: pathlib.Path,
    show: bool,
) -> None:
    """Горизонтальные столбики: доля «Да» / «Нет» для каждого критерия."""
    if not boolean_rows:
        return

    counts: dict[str, int] = {}
    totals: dict[str, int] = {}
    for key, _ in BOOLEAN_CRITERIA:
        values = [r[key] for r in boolean_rows if r[key] is not None]
        counts[key] = sum(1 for v in values if v)
        totals[key] = len(values)

    labels   = [label for _, label in BOOLEAN_CRITERIA]
    pass_pct = [counts[k] / totals[k] * 100 if totals[k] else 0 for k, _ in BOOLEAN_CRITERIA]
    fail_pct = [100 - p for p in pass_pct]

    fig, ax = plt.subplots(figsize=(9, 4))
    y = np.arange(len(labels))

    ax.barh(y, pass_pct, color="#2ecc71", label="Да", zorder=3)
    ax.barh(y, fail_pct, left=pass_pct, color="#e74c3c", label="Нет", zorder=3)

    for i, (p, (k, _)) in enumerate(zip(pass_pct, BOOLEAN_CRITERIA)):
        if p > 8:
            ax.text(p / 2, i, f"{p:.0f}%", ha="center", va="center",
                    color="white", fontweight="bold", fontsize=9)
        fp = 100 - p
        if fp > 8:
            ax.text(p + fp / 2, i, f"{fp:.0f}%", ha="center", va="center",
                    color="white", fontweight="bold", fontsize=9)
        ax.text(103, i, f"n={totals[k]}", va="center", fontsize=8, color="#555")

    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=9)
    ax.set_xlim(0, 115)
    ax.set_xlabel("Доля (%)")
    ax.set_title("Выполнение критериев оценки")
    ax.legend(loc="lower right", fontsize=9)
    ax.xaxis.grid(True, linestyle=":", alpha=0.4, zorder=0)
    ax.set_axisbelow(True)
    fig.tight_layout()
    _save(fig, out_dir, "boolean_criteria.png", show)


def plot_score_distribution(
    quality_rows: list[dict],
    out_dir: pathlib.Path,
    show: bool,
) -> None:
    """Гистограмма распределения всех числовых оценок."""
    all_scores = [
        r[m] for r in quality_rows for m in QUALITY_METRICS if r[m] is not None
    ]
    if not all_scores:
        return

    fig, ax = plt.subplots(figsize=(7, 4))

    for score in range(6):
        count = all_scores.count(score)
        ax.bar(score, count, color=SCORE_COLORS[score], edgecolor="white",
               linewidth=0.8, zorder=3)
        if count:
            ax.text(score, count + 0.4, str(count),
                    ha="center", va="bottom", fontsize=9)

    mean_val = float(np.mean(all_scores))
    ax.axvline(mean_val, color="navy", linestyle="--", linewidth=1.5,
               label=f"Среднее: {mean_val:.2f}")

    ax.set_xticks(range(6))
    ax.set_xlabel("Оценка")
    ax.set_ylabel("Количество")
    ax.set_title(f"Распределение оценок  (всего: {len(all_scores)})")
    ax.legend(fontsize=9)
    ax.yaxis.grid(True, linestyle=":", alpha=0.4, zorder=0)
    ax.set_axisbelow(True)
    fig.tight_layout()
    _save(fig, out_dir, "score_distribution.png", show)


# ---------------------------------------------------------------------------
# Графики в разрезе дисциплин
# ---------------------------------------------------------------------------

def plot_disc_quality_averages(
    quality_rows: list[dict],
    out_dir: pathlib.Path,
    show: bool,
) -> None:
    """Горизонтальные сгруппированные столбики: средние оценки по дисциплинам."""
    if not quality_rows:
        return

    disc_values: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))
    for r in quality_rows:
        for m in QUALITY_METRICS:
            if r[m] is not None:
                disc_values[r["discipline"]][m].append(r[m])

    disciplines = sorted(disc_values)
    n_disc = len(disciplines)
    n_metrics = len(QUALITY_METRICS)
    colors = ["#e74c3c", "#3498db", "#2ecc71", "#f39c12"]
    bar_h = 0.18

    fig_h = max(4.0, n_disc * n_metrics * bar_h + n_disc * 0.3 + 1.5)
    fig, ax = plt.subplots(figsize=(9, fig_h))

    y_base = np.arange(n_disc, dtype=float)

    for i, m in enumerate(QUALITY_METRICS):
        means = [
            float(np.mean(disc_values[d][m])) if disc_values[d][m] else 0.0
            for d in disciplines
        ]
        offset = (i - (n_metrics - 1) / 2) * bar_h
        bars = ax.barh(y_base + offset, means, bar_h,
                       label=METRIC_LABELS[m], color=colors[i], zorder=3)
        for bar, val in zip(bars, means):
            if val > 0.3:
                ax.text(
                    val + 0.05,
                    bar.get_y() + bar.get_height() / 2,
                    f"{val:.1f}",
                    va="center", fontsize=7,
                )

    ax.set_xlim(0, 5.9)
    ax.axvline(5, color="gray", linestyle="--", linewidth=0.8, alpha=0.5)
    ax.set_yticks(y_base)
    ax.set_yticklabels(disciplines, fontsize=8)
    ax.set_xlabel("Средняя оценка")
    ax.set_title("Средние оценки качества по дисциплинам")
    ax.legend(loc="lower right", fontsize=8)
    ax.xaxis.grid(True, linestyle=":", alpha=0.4, zorder=0)
    ax.set_axisbelow(True)
    fig.tight_layout()
    _save(fig, out_dir, "disc_quality_averages.png", show)


def plot_disc_boolean_criteria(
    boolean_rows: list[dict],
    out_dir: pathlib.Path,
    show: bool,
) -> None:
    """Тепловая карта: строки = дисциплины, столбцы = критерии, ячейки = % «Да»."""
    if not boolean_rows:
        return

    disciplines = sorted({r["discipline"] for r in boolean_rows})
    criteria_keys = [k for k, _ in BOOLEAN_CRITERIA]
    criteria_labels = [lbl.replace("\n", " ") for _, lbl in BOOLEAN_CRITERIA]

    data = np.full((len(disciplines), len(criteria_keys)), np.nan)
    for di, disc in enumerate(disciplines):
        for ci, key in enumerate(criteria_keys):
            vals = [r[key] for r in boolean_rows
                    if r["discipline"] == disc and r[key] is not None]
            if vals:
                data[di, ci] = sum(1 for v in vals if v) / len(vals) * 100

    fig_h = max(3.0, len(disciplines) * 0.55 + 1.5)
    fig, ax = plt.subplots(figsize=(10, fig_h))

    im = ax.imshow(data, cmap="RdYlGn", vmin=0, vmax=100, aspect="auto")

    ax.set_xticks(range(len(criteria_keys)))
    ax.set_xticklabels(criteria_labels, fontsize=9)
    ax.set_yticks(range(len(disciplines)))
    ax.set_yticklabels(disciplines, fontsize=8)
    ax.tick_params(top=True, labeltop=True, bottom=False, labelbottom=False)

    for di in range(len(disciplines)):
        for ci in range(len(criteria_keys)):
            val = data[di, ci]
            if not np.isnan(val):
                ax.text(ci, di, f"{val:.0f}%", ha="center", va="center",
                        fontsize=8, fontweight="bold",
                        color="white" if val < 30 or val > 85 else "black")

    plt.colorbar(im, ax=ax, label="Доля «Да» (%)", shrink=0.7)
    ax.set_title("Выполнение критериев по дисциплинам (%)", pad=14)
    fig.tight_layout()
    _save(fig, out_dir, "disc_boolean_criteria.png", show)


# ---------------------------------------------------------------------------
# Извлечение полных данных для Excel
# ---------------------------------------------------------------------------

def extract_full_results(
    all_data: list[dict],
) -> tuple[list[dict], list[dict]]:
    """Возвращает (disc_rows, lab_rows) со всеми полями из analysis.json."""
    disc_rows: list[dict] = []
    lab_rows:  list[dict] = []

    for run in all_data:
        run_name = run["_run_name"]
        for result in run.get("Results", []):
            test_id   = result.get("RunId", "?")
            disc      = result.get("Discipline", {})
            disc_name = normalize_discipline(disc.get("Name", ""))

            disc_rows.append({
                "Прогон":                    run_name,
                "Дисциплина":               disc_name,
                "Тест":                      test_id,
                "Соответствует дисциплине": disc.get("AssignmentsMatchDiscipline"),
                "Причина (соответствие)":   disc.get("AssignmentsMatchReason", ""),
                "ЛР различаются":           disc.get("LabsDiffer"),
                "Причина (различие ЛР)":    disc.get("LabsDifferReason", ""),
                "Последовательность логична": disc.get("SequenceLogical"),
                "Причина (последовательность)": disc.get("SequenceReason", ""),
            })

            quality_by_lab = {
                q.get("LabNumber"): q
                for q in result.get("Quality", [])
            }

            for lab in result.get("Labs", []):
                lab_n = lab.get("LabNumber")
                q     = quality_by_lab.get(lab_n, {})
                metrics = {m: q.get(m) for m in QUALITY_METRICS}
                vals    = [v for v in metrics.values() if v is not None]
                avg     = round(sum(vals) / len(vals), 2) if vals else None

                lab_rows.append({
                    "Прогон":               run_name,
                    "Дисциплина":           disc_name,
                    "Тест":                 test_id,
                    "Номер ЛР":             lab_n,
                    "Название ЛР":          lab.get("AssignmentTitle",
                                                    q.get("AssignmentTitle", "")),
                    "Варианты отличаются":  lab.get("VariantsDiffer"),
                    "Описание отличий":     lab.get("VariantsDifferences", ""),
                    "Одинаковая сложность": lab.get("VariantsSameDifficulty"),
                    "Причина (сложность)":  lab.get("VariantsDifficultyReason", ""),
                    "Ошибка генерации":     lab.get("MissingGenerationReason", ""),
                    "Корректность":         metrics.get("Correctness"),
                    "Качество":             metrics.get("Quality"),
                    "Полнота":              metrics.get("Completeness"),
                    "Ясность":              metrics.get("Clarity"),
                    "Среднее":              avg,
                    "Обоснование":          q.get("Justification", ""),
                })

    return disc_rows, lab_rows


# ---------------------------------------------------------------------------
# Экспорт в Excel
# ---------------------------------------------------------------------------

def _bool_str(val: bool | None) -> str:
    if val is True:  return "Да"
    if val is False: return "Нет"
    return ""


def _write_sheet_header(ws, headers: list[str], hdr_font, hdr_fill, hdr_align) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align


def _autofit(ws, min_w: int = 8, max_w: int = 55) -> None:
    for col in ws.columns:
        length = max(
            len(str(cell.value or "")) for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, min_w), max_w)


def _color_scale(ws, col_letter: str, n_rows: int, lo: str, mid: str, hi: str) -> None:
    from openpyxl.formatting.rule import ColorScaleRule
    rule = ColorScaleRule(
        start_type="num", start_value=0,   start_color=lo,
        mid_type="num",   mid_value=3,     mid_color=mid,
        end_type="num",   end_value=5,     end_color=hi,
    )
    ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{n_rows + 1}", rule)


def _bool_fills(ws, col_letter: str, n_rows: int,
                yes_fill, no_fill, rows: list[dict], key: str) -> None:
    from openpyxl.utils import column_index_from_string
    col_idx = column_index_from_string(col_letter)
    for row_i, row_data in enumerate(rows, start=2):
        cell = ws.cell(row=row_i, column=col_idx)
        if row_data.get(key) is True:
            cell.fill = yes_fill
        elif row_data.get(key) is False:
            cell.fill = no_fill


def export_excel(
    disc_rows: list[dict],
    lab_rows:  list[dict],
    quality_rows:  list[dict],
    boolean_rows:  list[dict],
    out_dir: pathlib.Path,
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [skip] Excel: openpyxl не установлен — выполните: pip install openpyxl")
        return

    out_dir.mkdir(parents=True, exist_ok=True)

    # Общие стили
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="2F5496")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    yes_fill  = PatternFill("solid", fgColor="C6EFCE")
    no_fill   = PatternFill("solid", fgColor="FFC7CE")
    num_align = Alignment(horizontal="center")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------------ #
    # Лист 1: «Качество ЛР»                                              #
    # ------------------------------------------------------------------ #
    ws = wb.create_sheet("Качество ЛР")
    score_cols = ["Корректность", "Качество", "Полнота", "Ясность", "Среднее"]
    headers = [
        "Прогон", "Дисциплина", "Тест", "Номер ЛР", "Название ЛР",
        *score_cols, "Обоснование",
    ]
    _write_sheet_header(ws, headers, hdr_font, hdr_fill, hdr_align)

    for r in lab_rows:
        ws.append([
            r["Прогон"], r["Дисциплина"], r["Тест"], r["Номер ЛР"], r["Название ЛР"],
            r["Корректность"], r["Качество"], r["Полнота"], r["Ясность"], r["Среднее"],
            r["Обоснование"],
        ])

    # Числовое выравнивание и цветовые шкалы для оценок
    score_col_letters = [get_column_letter(headers.index(h) + 1) for h in score_cols]
    for ltr in score_col_letters:
        for row_i in range(2, len(lab_rows) + 2):
            ws.cell(row=row_i, column=headers.index(score_cols[0]) + 1 + score_col_letters.index(ltr)).alignment = num_align
        _color_scale(ws, ltr, len(lab_rows), "FF0000", "FFFF00", "00B050")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autofit(ws)

    # ------------------------------------------------------------------ #
    # Лист 2: «Критерии дисциплин»                                       #
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet("Критерии дисциплин")
    bool_disc_keys = [
        "Соответствует дисциплине",
        "ЛР различаются",
        "Последовательность логична",
    ]
    h2 = [
        "Прогон", "Дисциплина", "Тест",
        "Соответствует дисциплине", "Причина (соответствие)",
        "ЛР различаются",           "Причина (различие ЛР)",
        "Последовательность логична", "Причина (последовательность)",
    ]
    _write_sheet_header(ws2, h2, hdr_font, hdr_fill, hdr_align)

    for r in disc_rows:
        ws2.append([
            r["Прогон"], r["Дисциплина"], r["Тест"],
            _bool_str(r["Соответствует дисциплине"]), r["Причина (соответствие)"],
            _bool_str(r["ЛР различаются"]),            r["Причина (различие ЛР)"],
            _bool_str(r["Последовательность логична"]), r["Причина (последовательность)"],
        ])

    for key, col in zip(bool_disc_keys, ["D", "F", "H"]):
        _bool_fills(ws2, col, len(disc_rows), yes_fill, no_fill, disc_rows, key)

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    _autofit(ws2)

    # ------------------------------------------------------------------ #
    # Лист 3: «Критерии ЛР»                                              #
    # ------------------------------------------------------------------ #
    ws3 = wb.create_sheet("Критерии ЛР")
    bool_lab_keys = ["Варианты отличаются", "Одинаковая сложность"]
    h3 = [
        "Прогон", "Дисциплина", "Тест", "Номер ЛР", "Название ЛР",
        "Варианты отличаются", "Описание отличий",
        "Одинаковая сложность", "Причина (сложность)",
        "Ошибка генерации",
    ]
    _write_sheet_header(ws3, h3, hdr_font, hdr_fill, hdr_align)

    for r in lab_rows:
        ws3.append([
            r["Прогон"], r["Дисциплина"], r["Тест"], r["Номер ЛР"], r["Название ЛР"],
            _bool_str(r["Варианты отличаются"]), r["Описание отличий"],
            _bool_str(r["Одинаковая сложность"]), r["Причина (сложность)"],
            r["Ошибка генерации"],
        ])

    for key, col in zip(bool_lab_keys, ["F", "H"]):
        _bool_fills(ws3, col, len(lab_rows), yes_fill, no_fill, lab_rows, key)

    # Ошибки генерации — выделить непустые
    err_fill = PatternFill("solid", fgColor="FFE699")
    for row_i, r in enumerate(lab_rows, start=2):
        if r["Ошибка генерации"]:
            ws3.cell(row=row_i, column=10).fill = err_fill

    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions
    _autofit(ws3)

    # ------------------------------------------------------------------ #
    # Лист 4: «Сводка по дисциплинам»                                    #
    # ------------------------------------------------------------------ #
    ws4 = wb.create_sheet("Сводка по дисциплинам")
    h4 = [
        "Дисциплина", "Кол-во тестов", "Кол-во ЛР",
        "Ср. Корректность", "Ср. Качество", "Ср. Полнота", "Ср. Ясность", "Ср. Общее",
        "% Соответствует дисциплине", "% ЛР различаются",
        "% Последовательность логична", "% Варианты отличаются", "% Одинаковая сложность",
    ]
    _write_sheet_header(ws4, h4, hdr_font, hdr_fill, hdr_align)

    disciplines = sorted({r["discipline"] for r in quality_rows})
    disc_summary_rows: list[dict] = []
    for disc in disciplines:
        q_disc = [r for r in quality_rows  if r["discipline"] == disc]
        b_disc = [r for r in boolean_rows  if r["discipline"] == disc]
        n_tests = len({r["test"] for r in q_disc})
        n_labs  = len(q_disc)

        def avg_metric(m: str) -> float | None:
            vals = [r[m] for r in q_disc if r[m] is not None]
            return round(sum(vals) / len(vals), 2) if vals else None

        def pct(key: str) -> float | None:
            vals = [r[key] for r in b_disc if r[key] is not None]
            return round(sum(1 for v in vals if v) / len(vals) * 100, 1) if vals else None

        all_scores = [r[m] for r in q_disc for m in QUALITY_METRICS if r[m] is not None]
        avg_total  = round(sum(all_scores) / len(all_scores), 2) if all_scores else None

        row = {
            "disc": disc, "n_tests": n_tests, "n_labs": n_labs,
            "Ср. Корректность": avg_metric("Correctness"),
            "Ср. Качество":     avg_metric("Quality"),
            "Ср. Полнота":      avg_metric("Completeness"),
            "Ср. Ясность":      avg_metric("Clarity"),
            "Ср. Общее":        avg_total,
            "% Соответствует дисциплине":   pct("AssignmentsMatchDiscipline"),
            "% ЛР различаются":             pct("LabsDiffer"),
            "% Последовательность логична": pct("SequenceLogical"),
            "% Варианты отличаются":        pct("VariantsDiffer"),
            "% Одинаковая сложность":       pct("VariantsSameDifficulty"),
        }
        disc_summary_rows.append(row)
        ws4.append([
            disc, n_tests, n_labs,
            row["Ср. Корректность"], row["Ср. Качество"],
            row["Ср. Полнота"],      row["Ср. Ясность"], avg_total,
            row["% Соответствует дисциплине"],   row["% ЛР различаются"],
            row["% Последовательность логична"], row["% Варианты отличаются"],
            row["% Одинаковая сложность"],
        ])

    n4 = len(disc_summary_rows)
    for col_i, ltr in enumerate(["D", "E", "F", "G", "H"], start=0):
        _color_scale(ws4, ltr, n4, "FF0000", "FFFF00", "00B050")

    ws4.freeze_panes = "A2"
    _autofit(ws4)

    # ------------------------------------------------------------------ #
    # Лист 5: «Сводка по прогонам»                                       #
    # ------------------------------------------------------------------ #
    ws5 = wb.create_sheet("Сводка по прогонам")
    h5 = [
        "Прогон", "Кол-во тестов", "Кол-во ЛР",
        "Ср. Корректность", "Ср. Качество", "Ср. Полнота", "Ср. Ясность", "Ср. Общее",
    ]
    _write_sheet_header(ws5, h5, hdr_font, hdr_fill, hdr_align)

    runs = sorted({r["run"] for r in quality_rows})
    n5 = len(runs)
    for run_name in runs:
        q_run = [r for r in quality_rows if r["run"] == run_name]
        n_tests = len({r["test"] for r in q_run})
        n_labs  = len(q_run)

        def avg_m(m: str) -> float | None:
            vals = [r[m] for r in q_run if r[m] is not None]
            return round(sum(vals) / len(vals), 2) if vals else None

        all_s  = [r[m] for r in q_run for m in QUALITY_METRICS if r[m] is not None]
        avg_t  = round(sum(all_s) / len(all_s), 2) if all_s else None
        ws5.append([
            run_name, n_tests, n_labs,
            avg_m("Correctness"), avg_m("Quality"),
            avg_m("Completeness"), avg_m("Clarity"), avg_t,
        ])

    for ltr in ["D", "E", "F", "G", "H"]:
        _color_scale(ws5, ltr, n5, "FF0000", "FFFF00", "00B050")

    ws5.freeze_panes = "A2"
    _autofit(ws5)

    # ------------------------------------------------------------------ #
    path = out_dir / "analysis.xlsx"
    wb.save(path)
    print(f"  сохранён: {path}")


# ---------------------------------------------------------------------------
# Сводка в консоль
# ---------------------------------------------------------------------------

def print_summary(
    all_data: list[dict],
    quality_rows: list[dict],
    boolean_rows: list[dict],
) -> None:
    total_tests = sum(len(d.get("Results", [])) for d in all_data)

    print(f"\n{'-'*48}")
    print("  Итоговая сводка")
    print(f"{'-'*48}")
    print(f"  Прогонов анализа : {len(all_data)}")
    print(f"  Тест-кейсов      : {total_tests}")
    print(f"  Записей ЛР       : {len(quality_rows)}")

    all_scores = [
        r[m] for r in quality_rows for m in QUALITY_METRICS if r[m] is not None
    ]
    if all_scores:
        print("\n  Качество (среднее / мин / макс):")
        for m in QUALITY_METRICS:
            vals = [r[m] for r in quality_rows if r[m] is not None]
            if vals:
                print(f"    {METRIC_LABELS[m]:<16}  {np.mean(vals):.2f}  "
                      f"(мин {min(vals)}, макс {max(vals)})")
        print(f"    {'Общее':<16}  {np.mean(all_scores):.2f}")

    bool_keys = [k for k, _ in BOOLEAN_CRITERIA]
    have_bool = any(
        r[k] is not None for r in boolean_rows for k in bool_keys
    )
    if have_bool:
        print("\n  Критерии (доля «Да»):")
        for key, label in BOOLEAN_CRITERIA:
            vals = [r[key] for r in boolean_rows if r[key] is not None]
            if vals:
                pct = sum(1 for v in vals if v) / len(vals) * 100
                clean = label.replace("\n", " ")
                print(f"    {clean:<32}  {pct:.0f}%  (n={len(vals)})")

    # --- В разрезе дисциплин ---
    disciplines = sorted({r["discipline"] for r in quality_rows})
    if len(disciplines) > 1:
        print(f"\n  По дисциплинам ({len(disciplines)} шт.):")
        for disc in disciplines:
            disc_scores = [
                r[m] for r in quality_rows
                if r["discipline"] == disc
                for m in QUALITY_METRICS
                if r[m] is not None
            ]
            n_tests = len({r["test"] for r in quality_rows if r["discipline"] == disc})
            mean_str = f"{np.mean(disc_scores):.2f}" if disc_scores else "—"
            short = disc if len(disc) <= 45 else disc[:42] + "..."
            print(f"    {short:<45}  среднее {mean_str}  (тестов: {n_tests})")
    print()


# ---------------------------------------------------------------------------
# Точка входа
# ---------------------------------------------------------------------------

def main() -> None:
    script_dir = pathlib.Path(__file__).resolve().parent
    default_candidates = [
        script_dir.parent.parent / "LabGenerator.TeacherEmulator" / "results" / "journal-analysis",
        script_dir.parent.parent / "artifacts" / "journal-analysis",
    ]
    default_input = next((p for p in default_candidates if p.exists()), default_candidates[0])

    parser = argparse.ArgumentParser(
        description="Визуализация результатов анализа журналов LabGenerator.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--input", "-i",
        default=str(default_input),
        metavar="PATH",
        help="Путь к каталогу с analysis.json или к самому файлу. "
             f"По умолчанию: {default_input}",
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        metavar="DIR",
        help="Каталог для сохранения графиков. "
             "По умолчанию: <input>/charts/ для каталога, "
             "<input parent>/charts/ для файла.",
    )
    parser.add_argument(
        "--show",
        action="store_true",
        help="Показать графики в интерактивном окне (дополнительно к сохранению).",
    )
    parser.add_argument(
        "--max-heatmap-rows",
        type=int,
        default=60,
        metavar="N",
        help="Максимальное число строк в тепловой карте (по умолчанию: 60).",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Не создавать analysis.xlsx.",
    )
    args = parser.parse_args()

    input_path = pathlib.Path(args.input)
    if not input_path.exists():
        print(f"[error] Путь не найден: {input_path}", file=sys.stderr)
        sys.exit(1)

    out_dir = pathlib.Path(args.output) if args.output else (
        input_path.parent / "charts" if input_path.is_file() else input_path / "charts"
    )

    print(f"Поиск analysis.json в: {input_path}")
    paths = find_analysis_files(input_path)
    if not paths:
        print("[error] Файлы analysis.json не найдены.", file=sys.stderr)
        sys.exit(1)
    print(f"Найдено файлов: {len(paths)}")

    all_data      = load_all(paths)
    quality_rows  = extract_quality_rows(all_data)
    boolean_rows  = extract_boolean_rows(all_data)
    disc_rows, lab_rows = extract_full_results(all_data)

    print(f"\nГенерация результатов -> {out_dir}")

    print("  [общий анализ]")
    plot_quality_heatmap(quality_rows, out_dir, args.show, args.max_heatmap_rows)
    plot_quality_averages(quality_rows, out_dir, args.show)
    plot_boolean_criteria(boolean_rows, out_dir, args.show)
    plot_score_distribution(quality_rows, out_dir, args.show)

    disciplines = sorted({r["discipline"] for r in quality_rows})
    if len(disciplines) > 1:
        print(f"  [анализ по дисциплинам: {len(disciplines)} шт.]")
        plot_disc_quality_averages(quality_rows, out_dir, args.show)
        plot_disc_boolean_criteria(boolean_rows, out_dir, args.show)
    else:
        print("  [анализ по дисциплинам пропущен — найдена только одна дисциплина]")

    if not args.no_excel:
        print("  [экспорт Excel]")
        export_excel(disc_rows, lab_rows, quality_rows, boolean_rows, out_dir)

    print_summary(all_data, quality_rows, boolean_rows)
    print(f"Готово. Результаты в: {out_dir}")


if __name__ == "__main__":
    main()
